"""Inventory export pipeline (READ-ONLY) — powers the taobao_export_inventory tool.

Drives the buyer order list's own pagination (the only path to full history), parses the
dinamic template, computes per-line **landed cost** (product price + order shipping allocated
by quantity), categorizes products, and writes a visual xlsx (embedded thumbnails, or =IMAGE
URLs for Google Sheets). Output + the order cache live under output/ (gitignored — order PII).

Why pagination, not an API loop: lib.mtop queryboughtlistv2 caps at the recent ~107 orders;
the page's own pager returns the dinamic format (data.data node map) with the whole history.
See the taobao-order-history-limit memory.
"""

from __future__ import annotations

import concurrent.futures
import hashlib
import io
import json
import os
import re
import urllib.request

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_BOUGHT_URL = "https://buyertrade.taobao.com/trade/itemlist/list_bought_items.htm"
_OUT_DIR = "output"
_CACHE = os.path.join(_OUT_DIR, ".inventory_orders.json")   # by_oid cache (gitignored)
THUMB = 60

# Click the footer 下一页 (Ant Design pager); returns whether it clicked + the active page number.
_NEXT_JS = r"""() => {
  const norm = s => (s || '').replace(/\s+/g, ' ').trim();
  let cur = null;
  const act = document.querySelector('.ant-pagination-item-active');
  if (act) cur = parseInt(norm(act.innerText)) || null;
  const nx = document.querySelector('.ant-pagination-next');
  if (!nx) return { clicked: false, reason: 'no-next', cur };
  const cls = (nx.className || '').toString();
  if (/ant-pagination-disabled/.test(cls) || nx.getAttribute('aria-disabled') === 'true')
    return { clicked: false, reason: 'disabled', cur };
  try { nx.scrollIntoView({ block: 'center' }); } catch (e) {}
  (nx.querySelector('button,a') || nx).click();
  return { clicked: true, cur };
}"""

_CUSTOM_WORDS = ("补差价", "差价", "定金", "尾款", "运费", "邮费", "专拍", "补拍", "改价", "补邮费")
_FOOD_LINE = ("商家配送", "餐盒")
_FOOD_SELLER = ("麦当劳", "麦乐送", "肯德基", "宅急送", "古茗", "煲珠公", "瑞幸", "星巴克", "必胜客",
                "奶茶", "咖啡", "烘焙", "蛋糕", "菜馆", "食堂", "外卖", "生鲜", "水果店", "茶饮")

# First-match-wins keyword → category (fallback when no prior categorization is recoverable).
_CAT_RULES = [
    (("逆变器", "ups", "纯正弦", "储能", "电瓶", "车载电源"), "Inverter / UPS / Battery"),
    (("18650", "锂电", "锂离子", "充电器", "充电头", "bms", "保护板", "电池盒"), "Battery / Charger"),
    (("墨仓", "墨盒", "喷墨", "连供", "ciss", "爱普生", "epson", "清洗液", "打印头", "l805", "l130", "l310", "r330"),
     "Printer / Ink / CISS"),
    (("3d打印", "小鲁班", "pla", "petg", "tpu", "耗材", "喷嘴", "热床", "打印线材"), "3D printer / Filament"),
    (("舵机", "电机", "马达", "步进", "空心杯", "螺旋桨", "齿轮", "联轴", "轴承", "servo", "motor"),
     "Motor / Servo / Mechanical"),
    (("esp32", "esp8266", "esp-01", "esp-05", "stm32", "arduino", "树莓", "raspberry", "开发板", "核心板",
      "单片机", "nano", "pico", "rp2040", "超级mini", "supermini"), "MCU / Dev board"),
    (("传感", "sensor", "陀螺", "加速度", "温湿度", "红外", "超声", "光敏", "霍尔", "mpu", "dht",
      "麦克风", "i2s", "音频", "气压", "距离模块", "wifi模块", "蓝牙模块", "透传"), "Sensor / Module"),
    (("dc-dc", "dcdc", "升压", "降压", "稳压模块", "电源模块", "buck", "boost", "mppt", "升降压"),
     "DC-DC / Power module"),
    (("杜邦", "端子", "连接器", "排针", "排母", "线束", "跳线", "数据线", "ghs", "jst", "母对母", "公对母",
      "公对公", "连接线", "1.25mm", "2.54", "xh2.54", "ph2.0"), "Connector / Wire / Cable"),
    (("电阻", "电容", "电感", "二极管", "晶振", "三极管", "mos管", "继电器", "保险丝", "磁珠", "稳压管"),
     "Passive (R/L/C/crystal/diode)"),
    (("oled", "lcd", "数码管", "点阵", "显示屏", "tft", "屏幕", "led灯", "灯珠"), "Display / LED"),
    (("万用表", "烙铁", "焊台", "示波", "内阻测试", "螺丝刀", "镊子", "热风枪", "测试仪", "表笔", "钳"),
     "Tool / Instrument"),
    (("外壳", "机箱", "面板", "pcb", "洞洞板", "万能板", "面包板", "散热", "风扇", "支架", "亚克力"),
     "PCB / Proto / Enclosure"),
    (("ic", "芯片", "贴片", "tssop", "qfn", "sop", "sot", "集成电路", "原装进口", "驱动ic", "稳压ic"),
     "IC / Chip"),
    (("vero moda", "针织", "上衣", "马甲", "卫衣", "衬衫", "连衣裙", "女装", "男装", "牛仔", "外套", "鞋"),
     "Clothing / Personal"),
]


def _to_num(s) -> float:
    try:
        return float(re.sub(r"[^\d.]", "", s or "") or 0)
    except Exception:
        return 0.0


def _qty(it) -> int:
    try:
        return max(1, int(it["qty"]))
    except (TypeError, ValueError):
        return 1


def _is_food(seller: str, titles: list[str]) -> bool:
    if any(any(w in t for w in _FOOD_LINE) for t in titles):
        return True
    return any(w in (seller or "") for w in _FOOD_SELLER)


def accumulate_dinamic(body: str, by_oid: dict) -> None:
    """Merge one page's dinamic XHR body into by_oid {orderId: {...}} (keyed by node-key suffix)."""
    try:
        d = json.loads(body)
    except Exception:
        a = body.find("{")
        try:
            d = json.loads(body[a:body.rfind("}") + 1])
        except Exception:
            return
    dd = ((d.get("data") or {}).get("data")) or {}
    if not isinstance(dd, dict):
        return
    for k, v in dd.items():
        if not isinstance(v, dict):
            continue
        f = v.get("fields") or {}
        if k.startswith("shopInfo_"):
            oid = f.get("orderId") or k.split("_", 1)[1]
            o = by_oid.setdefault(oid, {})
            o["createDay"] = f.get("createDay") or o.get("createDay")
            o["seller"] = f.get("sellerName") or o.get("seller") or "?"
        elif k.startswith("orderStatus_"):
            m = re.match(r"orderStatus_(\d+)", k)
            if m:
                o = by_oid.setdefault(m.group(1), {})
                o["status"] = f.get("subTitle") or f.get("title") or o.get("status", "")
        elif k.startswith("orderPayment_"):
            m = re.match(r"orderPayment_(\d+)", k)
            if m:
                o = by_oid.setdefault(m.group(1), {})
                af = f.get("actualFee") if isinstance(f.get("actualFee"), dict) else {}
                pf = f.get("pcPostFee") if isinstance(f.get("pcPostFee"), dict) else {}
                o["order_paid"] = af.get("value") or o.get("order_paid")
                o["shipping"] = pf.get("value") or o.get("shipping")
        elif k.startswith("orderItemInfo_"):
            m = re.match(r"orderItemInfo_(\d+)", k)
            if not m:
                continue
            it = f.get("item") or {}
            pi = it.get("priceInfo") or {}
            o = by_oid.setdefault(m.group(1), {})
            seen = o.setdefault("_seen", set())
            key = (it.get("itemId"), str(it.get("title") or ""), str(it.get("skuText") or ""))
            if key in seen:
                continue
            seen.add(key)
            o.setdefault("items", []).append({
                "title": str(it.get("title") or "").strip(),
                "variant": str(it.get("skuText") or "").strip(),
                "price": pi.get("actualTotalFee") or pi.get("promotion") or "",
                "qty": it.get("quantity"),
                "itemId": it.get("itemId"),
                "pic": it.get("pic") or "",
                "itemUrl": (it.get("itemUrl") or "").split("&mi_id=")[0],
            })


def inventory_rows(by_oid: dict, since: str) -> list[dict]:
    """Order map → deduped, date-filtered rows with landed cost (shipping spread by qty)."""
    rows = []
    for oid, o in by_oid.items():
        day = o.get("createDay")
        if (day or "") < since:
            continue
        seller = o.get("seller") or "?"
        items = o.get("items") or []
        titles = [it["title"] for it in items]
        kind = "food/local" if _is_food(seller, titles) else "goods"
        total_units = sum(_qty(it) for it in items) or 1
        ship_per_unit = _to_num(o.get("shipping")) / total_units
        for it in items:
            title, variant = it["title"], it["variant"]
            qty = _qty(it)
            unit = _to_num(it["price"])
            line_total = round(unit * qty, 2)
            ship_line = round(ship_per_unit * qty, 2)
            landed_unit = round(unit + ship_per_unit, 2)
            landed_line = round(line_total + ship_line, 2)
            # Flag opaque payment-link lines (1元补差价 etc.) — decode the real product from the
            # vendor chat. Kept generic: no buyer-specific product mappings live in the tool.
            custom = any(w in title for w in _CUSTOM_WORDS) or (unit < 2 and qty >= 300)
            rows.append({
                "date": day, "seller": seller, "title": title, "variant": variant, "qty": qty,
                "unit": unit, "line_total": line_total, "ship": ship_line, "landed_unit": landed_unit,
                "landed_line": landed_line, "status": o.get("status") or "", "order_no": oid,
                "custom": custom, "kind": kind, "pic": it.get("pic") or "",
                "item_url": it.get("itemUrl") or "", "item_id": str(it.get("itemId") or ""),
            })
    seen, uniq = set(), []
    for r in rows:
        k = (r["order_no"], r["title"], r["variant"], r["line_total"])
        if k not in seen:
            seen.add(k)
            uniq.append(r)
    uniq.sort(key=lambda r: (r["date"] or "", str(r["order_no"])), reverse=True)
    return uniq


def _keyword_category(title: str) -> str:
    t = (title or "").lower()
    for words, cat in _CAT_RULES:
        if any(w in t for w in words):
            return cat
    return "Other"


def categorize(rows: list[dict], prior_xlsx: str) -> None:
    """Fill r['category'] in place. Prefer categories recovered from a prior inventory xlsx
    (so a good LLM categorization is preserved), else a keyword heuristic."""
    by_title = {}
    if prior_xlsx and os.path.exists(prior_xlsx):
        try:
            wb = load_workbook(prior_xlsx, read_only=True)
            ws = wb["Inventory"]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "Product" in header and "Category" in header:
                pi, ci = header.index("Product"), header.index("Category")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[pi] and row[ci]:
                        by_title[str(row[pi])] = row[ci]
            wb.close()
        except Exception:
            by_title = {}
    for r in rows:
        r["category"] = by_title.get(r["title"]) or _keyword_category(r["title"])


# ── images ──────────────────────────────────────────────────────────────────
def _img_path(url: str) -> str:
    return os.path.join(_OUT_DIR, ".inv_images", hashlib.md5(url.encode()).hexdigest()[:16] + ".png")


def _download_one(url: str):
    if not url:
        return None
    path = _img_path(url)
    if os.path.exists(path):
        return path
    full = ("https:" + url) if url.startswith("//") else url
    try:
        from PIL import Image as PILImage
        req = urllib.request.Request(full, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Referer": "https://buyertrade.taobao.com/",
        })
        data = urllib.request.urlopen(req, timeout=20).read()
        im = PILImage.open(io.BytesIO(data)).convert("RGB")
        im.thumbnail((THUMB, THUMB))
        os.makedirs(os.path.dirname(path), exist_ok=True)
        im.save(path, "PNG")
        return path
    except Exception:
        return None


def _download_images(urls: list[str]) -> dict:
    paths = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=16) as ex:
        futs = {ex.submit(_download_one, u): u for u in urls}
        for fu in concurrent.futures.as_completed(futs):
            p = fu.result()
            if p:
                paths[futs[fu]] = p
    return paths


# ── xlsx ────────────────────────────────────────────────────────────────────
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="2F5496")
_FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")
_MONEY = "¥#,##0.00"
_MONEY_HEADERS = ("Unit ¥", "Line ¥", "Ship ¥", "Landed/u ¥", "Landed ¥")


def build_xlsx(rows: list[dict], path: str, embed_images: bool) -> dict:
    """Write the inventory workbook. embed_images=True embeds thumbnails (Numbers/Excel);
    False writes =IMAGE() formulas (Google Sheets). Returns a summary dict."""
    paths = _download_images(list({r["pic"] for r in rows if r.get("pic")})) if embed_images else {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    cols = ["Image", "Date", "Category", "Seller", "Product", "Variant", "Qty", "Unit ¥",
            "Line ¥", "Ship ¥", "Landed/u ¥", "Landed ¥", "Status", "Order #"]
    if not embed_images:
        cols.append("Product link")
    cols.append("Flag")
    money_idx = [cols.index(h) + 1 for h in _MONEY_HEADERS]
    order_idx = cols.index("Order #") + 1
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).font = _HEAD_FONT
        ws.cell(1, c).fill = _HEAD_FILL
        ws.cell(1, c).alignment = Alignment(vertical="center", horizontal="center")

    n_img = 0
    for r in rows:
        url = (("https:" + r["pic"]) if r.get("pic", "").startswith("//") else r.get("pic", ""))
        image_cell = "" if embed_images else (f'=IMAGE("{url}")' if url else "")
        row = ["" if embed_images else image_cell, r["date"], r.get("category", ""), r["seller"],
               r["title"], r["variant"], r["qty"], r["unit"], r["line_total"], r["ship"],
               r["landed_unit"], r["landed_line"], r["status"], str(r["order_no"])]
        if not embed_images:
            link = r.get("item_url") or ""
            row.append(f'=HYPERLINK("{link}","open")' if link else "")
        row.append("custom-link — decode from vendor chat" if r["custom"] else "")
        ws.append(row)
        rn = ws.max_row
        ws.row_dimensions[rn].height = 46
        for ci in money_idx:
            ws.cell(rn, ci).number_format = _MONEY
        ws.cell(rn, order_idx).number_format = "@"
        for c in (5, 6):
            ws.cell(rn, c).alignment = Alignment(vertical="center", wrap_text=True)
        if r["custom"]:
            for c in range(1, len(cols) + 1):
                ws.cell(rn, c).fill = _FLAG_FILL
        if embed_images:
            p = paths.get(r.get("pic"))
            if p:
                img = XLImage(p)
                img.width = img.height = THUMB
                ws.add_image(img, f"A{rn}")
                n_img += 1
        elif url:
            n_img += 1

    widths = [13, 11, 22, 24, 42, 24, 5, 9, 10, 9, 11, 11, 12, 22] + ([14] if not embed_images else []) + [24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # By Category — landed spend
    from collections import defaultdict
    agg = defaultdict(lambda: [0, 0.0])
    for r in rows:
        c = r.get("category") or "Uncategorized"
        agg[c][0] += 1
        agg[c][1] += r["landed_line"]
    sm = wb.create_sheet("By Category")
    sm.append(["Category", "Lines", "Landed spend ¥"])
    for c in (1, 2, 3):
        sm.cell(1, c).font = _HEAD_FONT
        sm.cell(1, c).fill = _HEAD_FILL
    for cat, (n, sp) in sorted(agg.items(), key=lambda kv: -kv[1][1]):
        sm.append([cat, n, round(sp, 2)])
        sm.cell(sm.max_row, 3).number_format = _MONEY
    sm.append(["TOTAL", sum(n for n, _ in agg.values()), round(sum(sp for _, sp in agg.values()), 2)])
    sm.cell(sm.max_row, 1).font = Font(bold=True)
    sm.cell(sm.max_row, 3).number_format = _MONEY
    sm.column_dimensions["A"].width = 32
    sm.column_dimensions["C"].width = 16

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    dates = sorted(r["date"] for r in rows if r["date"])
    return {
        "path": path, "lines": len(rows), "orders": len({r["order_no"] for r in rows}),
        "landed_total": round(sum(r["landed_line"] for r in rows), 2),
        "date_range": f"{dates[0]} → {dates[-1]}" if dates else "—",
        "images": n_img, "flagged": sum(1 for r in rows if r["custom"]),
    }


# ── crawl ───────────────────────────────────────────────────────────────────
async def crawl_orders(since: str, max_clicks: int = 64) -> dict:
    """Page the buyer order list back until orders predate `since`; return {orderId: {...}}.
    Uses the warm session's single tab, paced + captcha-guarded (never bursts tabs)."""
    from src.browser.pacing import human_delay
    from src.browser.session import get_session

    session = get_session()
    page = await session.start()
    bodies: list[str] = []

    async def on_resp(resp):
        if "queryboughtlist" in resp.url.lower():
            try:
                bodies.append(await resp.text())
            except Exception:
                pass

    page.on("response", on_resp)
    await page.goto(_BOUGHT_URL, wait_until="domcontentloaded")
    await session.guard_captcha(page)
    await human_delay(4, 5)
    await page.mouse.wheel(0, 3200)
    await human_delay(1.5, 2.5)

    by_oid: dict = {}

    def drain():
        for b in bodies:
            accumulate_dinamic(b, by_oid)
        bodies.clear()

    def earliest():
        ds = [o.get("createDay") for o in by_oid.values() if o.get("createDay")]
        return min(ds) if ds else None

    drain()
    nogrow = 0
    for _ in range(max_clicks):
        e = earliest()
        if e and e < since:
            break
        before = len(by_oid)
        r = await page.evaluate(_NEXT_JS)
        if not r.get("clicked"):
            break
        for _ in range(6):
            await human_delay(1.0, 1.6)
            drain()
            if len(by_oid) > before:
                break
        await session.guard_captcha(page)
        await page.mouse.wheel(0, 3200)
        nogrow = nogrow + 1 if len(by_oid) == before else 0
        if nogrow >= 3:
            break
    try:
        page.remove_listener("response", on_resp)
    except Exception:
        pass
    for o in by_oid.values():
        o.pop("_seen", None)
    return by_oid


async def export_inventory(since: str = "2025-01-01", filename: str = "inventory_2025_2026.xlsx",
                           embed_images: bool = True, refresh: bool = True) -> dict:
    """Crawl (or reuse cache) → landed-cost rows → categorize → write workbook. Returns a summary.
    refresh=False reuses the last crawl cache (output/.inventory_orders.json) — no Taobao traffic."""
    os.makedirs(_OUT_DIR, exist_ok=True)
    if not refresh and os.path.exists(_CACHE):
        by_oid = json.load(open(_CACHE))
    else:
        by_oid = await crawl_orders(since)
        with open(_CACHE, "w") as f:
            json.dump(by_oid, f, ensure_ascii=False)

    rows = inventory_rows(by_oid, since)
    out_path = filename if os.path.isabs(filename) or filename.startswith(_OUT_DIR) \
        else os.path.join(_OUT_DIR, filename)
    categorize(rows, prior_xlsx=out_path)   # recover categories from a prior run if present
    import anyio
    return await anyio.to_thread.run_sync(build_xlsx, rows, out_path, embed_images)
