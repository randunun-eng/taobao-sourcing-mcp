"""Phase 4 tests: xlsx writer + markdown."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from src.extract.product import parse_product_res
from src.extract.reviews import group_by_variant
from src.models import Product, Review, SkuVariant
from src.output.markdown import product_to_markdown
from src.output.xlsx_writer import write_xlsx

FIXTURES = Path(__file__).parent / "fixtures"


def _p100() -> Product:
    res = json.loads((FIXTURES / "736546459871" / "detail_res.json").read_text(encoding="utf-8"))
    p = parse_product_res(res, "736546459871")
    p.reviews = [
        Review(rating=None, text="速度很快，插上就能用", has_images=True,
               sku_bought="P100 质保3年 以换代修", date="2026-05-26"),
        Review(rating=None, text="ollama和comfyui都ok", has_images=False,
               sku_bought="P100 质保7天 走量商品 退货承担运费20元", date="2026-03-19"),
    ]
    p.reviews_by_variant = group_by_variant(p.reviews)
    return p


def _col(ws, header: str) -> str:
    for c in ws[1]:
        if c.value == header:
            return c.column_letter
    raise AssertionError(f"header {header!r} not found")


def test_write_xlsx_opens_and_has_sheets(tmp_path):
    path = write_xlsx([_p100()], "cmp.xlsx", out_dir=str(tmp_path))
    assert Path(path).exists()
    wb = load_workbook(path)
    assert wb.sheetnames == ["Summary", "Variants", "Reviews"]


def test_variants_sheet_prices(tmp_path):
    path = write_xlsx([_p100()], "cmp.xlsx", out_dir=str(tmp_path))
    wb = load_workbook(path)
    wsv = wb["Variants"]
    pcol = _col(wsv, "Price ¥")
    prices = {wsv[f"{pcol}{r}"].value for r in range(2, 2 + 3)}
    assert prices == {420.0, 450.0, 400.0}


def test_summary_minmax(tmp_path):
    path = write_xlsx([_p100()], "cmp.xlsx", out_dir=str(tmp_path))
    wb = load_workbook(path)
    ws = wb["Summary"]
    assert ws[f"{_col(ws, 'Min ¥')}2"].value == 400.0
    assert ws[f"{_col(ws, 'Max ¥')}2"].value == 450.0


def test_reviews_sheet_rows(tmp_path):
    path = write_xlsx([_p100()], "cmp.xlsx", out_dir=str(tmp_path))
    wb = load_workbook(path)
    assert wb["Reviews"].max_row == 1 + 2   # header + 2 reviews


def test_no_crash_on_empty_product(tmp_path):
    p = Product(
        product_id="1", url="https://item.taobao.com/item.htm?id=1", title="单变体无评价",
        shop_name="x", price_range=(99.0, 99.0),
        variants=[SkuVariant(sku_id="s1", properties={"颜色": "黑"}, price=99.0, stock=3, available=True)],
        scraped_at="2026-06-03T00:00:00Z",
    )
    path = write_xlsx([p], "single.xlsx", out_dir=str(tmp_path))
    assert Path(path).exists()
    assert "单变体无评价" in product_to_markdown(p)


def test_markdown_has_prices():
    md = product_to_markdown(_p100())
    assert "特斯拉P100" in md or "P100" in md
    assert "¥420" in md and "¥450" in md
    assert "Reviews:" in md
    assert "Subsidy" in md   # the 平台加补后 vs 优惠前 caveat is surfaced


def test_variant_review_count_multigroup(tmp_path):
    """The Variants sheet '#Reviews(variant)' must be correct for a 颜色×尺寸 product."""
    from src.extract.reviews import group_by_variant

    variants = [
        SkuVariant(sku_id=f"{c}{s}", properties={"颜色": c, "尺寸": s}, price=99.0, stock=5, available=True)
        for c in ("黑", "白") for s in ("S", "L")
    ]
    reviews = [
        Review(rating=None, text="t1", has_images=True, sku_bought="黑 S", date="2026-01-01"),
        Review(rating=None, text="t2", has_images=False, sku_bought="黑 S", date="2026-01-02"),
        Review(rating=None, text="t3", has_images=False, sku_bought="白 L", date="2026-01-03"),
    ]
    p = Product(
        product_id="9", url="u", title="t", shop_name="s", price_range=(99.0, 99.0),
        variants=variants, reviews=reviews, reviews_by_variant=group_by_variant(reviews),
        scraped_at="2026-06-04T00:00:00Z",
    )
    wb = load_workbook(write_xlsx([p], "mg.xlsx", out_dir=str(tmp_path)))
    wsv = wb["Variants"]
    cc, sc, rc = _col(wsv, "颜色"), _col(wsv, "尺寸"), _col(wsv, "#Reviews(variant)")
    counts = {}
    for row in range(2, 2 + len(variants)):
        counts[f"{wsv[f'{cc}{row}'].value} {wsv[f'{sc}{row}'].value}"] = wsv[f"{rc}{row}"].value
    assert counts["黑 S"] == 2
    assert counts["白 L"] == 1
    assert counts["黑 L"] == 0   # the bug returned non-zero / wrong here


def test_variant_review_count_no_cross_contamination(tmp_path):
    """A '黑' variant must NOT inherit reviews keyed '黑色升级版' (substring bug)."""
    from src.extract.reviews import group_by_variant

    variants = [
        SkuVariant(sku_id="a", properties={"颜色": "黑"}, price=10.0, stock=1, available=True),
        SkuVariant(sku_id="b", properties={"颜色": "黑色升级版"}, price=20.0, stock=1, available=True),
    ]
    reviews = [Review(rating=None, text="t", has_images=False, sku_bought="黑色升级版", date="2026-01-01")]
    p = Product(
        product_id="9", url="u", title="t", shop_name="s", price_range=(10.0, 20.0),
        variants=variants, reviews=reviews, reviews_by_variant=group_by_variant(reviews),
        scraped_at="2026-06-04T00:00:00Z",
    )
    wb = load_workbook(write_xlsx([p], "cc.xlsx", out_dir=str(tmp_path)))
    wsv = wb["Variants"]
    cc, rc = _col(wsv, "颜色"), _col(wsv, "#Reviews(variant)")
    counts = {wsv[f"{cc}{r}"].value: wsv[f"{rc}{r}"].value for r in range(2, 4)}
    assert counts["黑"] == 0            # must NOT count the 黑色升级版 review
    assert counts["黑色升级版"] == 1
